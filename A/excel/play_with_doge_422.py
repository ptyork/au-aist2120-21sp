import openpyxl as xl
import shutil

def write(text):
    print(text, end='')

# MAYBE make a backup...
# shutil.copyfile('doge.xlsx', 'doge.xlsx.bak')

wb = xl.load_workbook('doge.xlsx')

# print(wb)

print(wb.sheetnames)

ws = wb["DogSheet"]
print(ws)

names = wb.sheetnames
ws = wb[names[0]]
print(ws)

ws = wb.worksheets[0]
print(ws)

cell = ws["A1"]
print(cell)
print(cell.value)

cell.value = "PAUL WUZ HERE"

# BADDDDDDD
# cell = "YEAH PAUL WUZ HERE AGAIN"
# print(cell)

cell = ws["A1"]
print(cell.value)

wb.save('doge.xlsx')



range = ws["C5:C8"]
print(range)
range = ws["D5:G9"]
print(range)
range = ws["D5:D9"]
print(range)

print(ws.rows)

print("ROWS")
for row in ws.rows:
    print(row)

print(ws.columns)

print("COLS")
for col in ws.columns:
    print(col)


prods = ws["C5:C8"] # a range is ALWAYS tuple of rows, each row a tuple of cells
for row in prods:
    for cell in row:
        write(cell.value + '\t')
print()

sales = ws["D5:G9"]
for row in sales:
    for cell in row:
        write(str(cell.value) + '\t')
    print()


col_letters = 'DEFG'
row_numbers = '5678'

for let in col_letters:
    # print(let)
    tot = 0
    for num in row_numbers:
        # print(num)
        # print(let+num)
        # OOPS!! NEED TO CHECK FOR A "NONE" VALUE
        addr = let+num
        val = ws[addr].value
        tot += val
    avg = tot / 4
    print(f"{let}: {avg}")

