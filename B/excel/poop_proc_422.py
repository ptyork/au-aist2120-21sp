import openpyxl as xl

def write(text):
    print(text, end='')

filename = 'poop.xlsx'

# HOW TO OPEN AN EXISTING EXCEL WORKBOOK
wb_poop = xl.load_workbook(filename)

# HOW TO CREATE A NEW EXCEL WORKBOOK
wb_new = xl.Workbook()

print(wb_poop)
print(wb_new)

wb = wb_poop  # COPY REFERENCE to a new variable name 'cause I'm too lazy to type wb_poop
print(wb)

# WORKBOOKS manage WORKSHEETS
print(wb.sheetnames)

# Get worksheet by name
print(wb['AwwSheet'])

# Get worksheet by position
print(wb.worksheets[0])

# Get worksheet stupidly
print(wb[wb.sheetnames[0]])

# Get the ACTIVE worksheet
print(wb.active)

ws = wb['AwwSheet']

# HOW TO ACCESS CELLS
# By NAME (ADDRESS)
print(ws["A1"])
print(ws["A1"].value)

# ALTER VALUE BY ADDRESS
ws["a1"].value = "PAUL WAS HERE"

print(ws["A1"].value)

# MAKE SURE TO USE THE RIGHT DATA TYPE
ws["D4"].value = 999.99

# JUST GET THE FORMULA, NOT THE CALCULATE VALUE
print(ws["D8"].value)

# SO LET'S CALCULATE A SUM

print(ws["D4:D7"])
print(ws["D4:G4"])
print(ws["D4:G7"])

range = ws["D4:G7"]

for row in range:
    for cell in row:
        # TRUE(ish) aka Truthy if string isn't empty, list isn't empty, number not zero, not Nne
        # OTHERWISE FALSEY
        if cell.value:    # TRUTHY
            write(f"{cell.value:>10}")
        else:
            write(f"{'N/A':>10}")
    print()

# YORK TRICK for iterating over Excel cells

col_string = 'DEFG'
row_string = '4567'

for let in col_string:
    # print(let)
    subtotal = 0
    for num in row_string:
        # print(num)
        # print(let+num)
        val = ws[let+num].value
        if val:   # TRUTHY
            subtotal += val

    print(f'{let}: {subtotal}')


wb.save(filename)
